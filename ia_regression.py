'''
Title: OLS Polynomial Regression in Python with Output to Excel and Plots
Created on Wed Mar 18 16:10:53 2015

@author: Eric M. Young
@email: ericyoung7@gmail.com

Version: Python 3.5.2
'''
import pandas as pd
import numpy as np
import statsmodels.api as sm
import statsmodels.stats.api as sms
from statsmodels.formula.api import ols
import matplotlib.pyplot as plt
import matplotlib.gridspec as grsp
from openpyxl import load_workbook
from openpyxl import Workbook

# Turn interactive plotting off
# NECESSARY TO PREVENT SQUISHING OF LARGER THAN SCREEN FIGURES!!!
plt.ioff()

def main() :
    
    lib_nomL = ['libA_avg', 'libAR_avg', 'libAF_avg', 
                'libB_avg', 'libC_avg', 'all_avg', 'highcad_avg']
    
    regL = ['1', '2', 'in']

    respL = ['ia', 'mmia', 'tia', 'sa', 'ma', 'ia2mmia']

    for n,x in enumerate(lib_nomL):
        for m,y in enumerate(regL):
            for l,z in enumerate(respL):
                
                filename = '%(r)s_%(o)s_%(t)s_regression.xlsx' % {'r':x, 
                                                                  'o': y, 
                                                                  't':z}
    
                # Load data to be fitted from Excel
                df = load_df(x)
                

                results = ols_regression(df, y, z)

                
                create_workbook(filename)
                fsL = fit_stats(results, filename)
                equation(results, filename)
                anova(results, filename)
                dbetas(results, filename)
                output_to_regression_history(fsL,x,y)
                plot_vars(results,x,y,z)
                plot_diag(results,x,y,z)
                #save_summary(results, filename)

                #Print select regression diagnostics data
                #print(results.summary())
                #print('Data Set:', regression_name)
                #print('Order:', regression_order)
                #print('# Observations:', results.nobs)
                #print('R Squared:', results.rsquared)
                #print('Cond. #:', fsL[2])
    
def create_workbook(f):
    
    wb = Workbook()
    
    # this is needed to rename the default first sheet
    ws = wb.active
    ws.title = 'equation'

    wb.save(f)

def load_df(lib):
    
    filename = "ia_regression_data" 
    worksheet = lib

    df = pd.read_excel('%(f)s.xlsx' % {'f':filename}, worksheet)
    
    return df

def ols_regression(df, reg, resp):
    
    respD = {'ia'  : 'AVG_IA',
             'mmia': 'AVG_MMIA',
             'tia' : 'AVG_TIA',
             'sa'  : 'AVG_SA',
             'ma'  : 'AVG_MA',
             'ia2mmia' : 'IA_to_MMIA'}
    
    #print(reg,resp)
    
    #General form for polynomial fits.  Specify the order above
    #Specify C for Categorical variables (i.e. Y/N) in the formula
    poly_modelsD = {'1' : '''%(r)s ~ (log_ITE + log_CAD + log_ACO + log_PYC + 
                            log_CSC + log_ACDH) ** 1''' % {'r':respD[resp]},
    
                    '2' : '''%(r)s ~ (log_ITE + log_CAD + log_ACO + log_PYC + 
                    log_CSC + log_ACDH) ** 2 + np.power(log_CAD,2) + 
                    np.power(log_ACO,2) + np.power(log_ITE,2) + 
                    np.power(log_PYC,2) + np.power(log_CSC,2) + 
                    np.power(log_ACDH,2)''' % {'r':respD[resp]},
                            
                    'in': '''%(r)s ~ (log_ITE + log_CAD + log_ACO + log_PYC + 
                    log_CSC + log_ACDH) ** 2''' % {'r':respD[resp]},
                              
                    'cad': '''%(r)s ~ (log_ITE + log_ACO + log_PYC + log_CSC + 
                    log_ACDH) ** 1''' % {'r':respD[resp]},
                               
                    '3' : '''%(r)s ~ (log_ITE + log_CAD + log_ACO + log_PYC + 
                    log_CSC + log_ACDH) ** 3 + np.power(log_CAD,2) + 
                    np.power(log_ACO,2) + np.power(log_ITE,2) + 
                    np.power(log_PYC,2) + np.power(log_CSC,2) + 
                    np.power(log_ACDH,2)+ np.power(log_CAD,3) + 
                    np.power(log_ACO,3) + np.power(log_ITE,3) + 
                    np.power(log_PYC,3) + np.power(log_CSC,3) + 
                    np.power(log_ACDH,3)''' % {'r':respD[resp]},
                    '4' : '''%(r)s ~ (log_ITE + log_CAD + log_ACO + log_PYC + 
                    log_CSC + log_ACDH) ** 4 + np.power(log_CAD,2) + 
                    np.power(log_ACO,2) + np.power(log_ITE,2) + 
                    np.power(log_PYC,2) + np.power(log_CSC,2) + 
                    np.power(log_ACDH,2)+ np.power(log_CAD,3) + 
                    np.power(log_ACO,3) + np.power(log_ITE,3) + 
                    np.power(log_PYC,3) + np.power(log_CSC,3) + 
                    np.power(log_ACDH,3)+ np.power(log_CAD,4) + 
                    np.power(log_ACO,4) + np.power(log_ITE,4) + 
                    np.power(log_PYC,4) + np.power(log_CSC,4) + 
                    np.power(log_ACDH,4)''' % {'r':respD[resp]}}

    regression=ols(poly_modelsD[reg],data=df).fit()

    return regression
    
def equation(res, f):
    
    ef = pd.DataFrame(res.params, columns=['coef'])
    ef['t'] = res.tvalues
    ef['p'] = res.pvalues
    ef['se'] = res.bse
    ef['HC0_se'] = res.HC0_se
    ef['HC1_se'] = res.HC1_se
    ef['HC2_se'] = res.HC2_se
    ef['HC3_se'] = res.HC3_se

    output_dataframe(ef, f, 'equation')

def fit_stats(res, f):
    
    # Diagnostics that require calculation (i.e. not in res.'')
    cn  = np.linalg.cond(res.model.exog)
    jb  = sms.jarque_bera(res.resid)[0]
    pjb = sms.jarque_bera(res.resid)[1]
    sk  = sms.jarque_bera(res.resid)[2]
    ku  = sms.jarque_bera(res.resid)[3]
    omni  = sms.omni_normtest(res.resid)[0]
    pomni = sms.omni_normtest(res.resid)[1]
    
    ff = pd.DataFrame([1], columns=['R Squared'])
    ff['R Squared'] = res.rsquared
    ff['Adj R Squared'] = res.rsquared_adj
    ff['# Observations'] = res.nobs
    ff['F Value'] = res.fvalue
    ff['Prob(F Value)'] = res.f_pvalue
    ff['Log-Likelihood'] = res.llf
    ff['Cond. #'] = cn
    ff['Jarque-Bera'] = jb
    ff['Prob(JB)'] = pjb
    ff['Skewness'] = sk
    ff['Kurtosis'] = ku
    ff['Omnibus'] = omni
    ff['Prob(Omnibus)'] = pomni
    ff['AIC'] = res.aic
    ff['BIC'] = res.bic
    
    output_dataframe(ff, f, 'fit')
    
    return [res.nobs, res.rsquared, cn, res.fvalue, res.f_pvalue, res.llf, sk, ku, omni, pomni]
    
def anova(res, f):

    ava = sm.stats.anova_lm(res, typ=2)
    
    output_dataframe(ava, f, 'anova')

def dbetas(res, f):
    
    infl = res.get_influence()

    sf = infl.summary_frame().filter(regex='dfb')
    
    sf['influential_observation_threshhold'] = 2./res.nobs**.5

    output_dataframe(sf, f, 'dfb')

def plot_vars(res,x,y,z):
    
    plt.figure(figsize=(4,10))
    # Use gridspec method to have different sized subplots for the bar and the line plots
    gs = grsp.GridSpec(6, 2, width_ratios=[1,1])

    varL = ['log_ITE', 'log_CAD', 'log_ACO', 'log_PYC', 'log_CSC', 'log_ACDH']

    pL = [2,2.25,2.5,2.75,3,3.25,3.5,3.75,4]

    cD = {'L': [2,2,2,2,2,2,2,2,2],
          'M': [3,3,3,3,3,3,3,3,3],
          'H': [4,4,4,4,4,4,4,4,4]}

    lD = {'L': 'b-', 'M': 'g-', 'H': 'r-'}

    for n, v in enumerate(varL) :
    
        ax = plt.subplot(gs[n])
    
        ax.set_title(v)
    
        for k in cD :
    
            pD = {'log_ITE':cD[k],'log_CAD':cD[k],'log_ACO':cD[k],'log_PYC':cD[k],'log_CSC':cD[k],'log_ACDH':cD[k]}
            
            pD[v] = pL

            p = res.predict(pD)
        
            plt.plot(pL,p,lD[k])
      
    plt.tight_layout(pad=2)
    plt.savefig('%(r)s_%(o)s_%(t)s_regression.pdf' % {'r': x, 
                                                      'o': y,
                                                      't': z})

def plot_diag(res,x,y,z):
    
    varL = ['log_ITE', 'log_CAD', 'log_ACO', 'log_PYC', 'log_CSC', 'log_ACDH']

    for n, v in enumerate(varL) :

        fig = sm.graphics.plot_regress_exog(res, n+1)
        fig.set_size_inches(10, 6)
        fig.savefig('%(r)s_%(o)s_%(t)s_%(v)s_diag.pdf' % {'r': x, 
                                                          'o': y, 
                                                          't': z,
                                                          'v': v})       
def output_to_regression_history(statsL, x, y):
    
    #Assemble data to be written to file
    new_row = [x, y]

    for n, x in enumerate(statsL):
        new_row.append(statsL[n])

    #Open history workbook
    his_book = load_workbook('ia_regression_history.xlsx')
    ws = his_book.get_sheet_by_name('reg_history')
    
    ws.append(new_row)
    
    his_book.save('ia_regression_history.xlsx')

def save_summary(res, f):
    
    print(res.summary())
    
    with open('%(f)s_summary.txt' % {'f':f}, 'w') as f:
        print(res.summary(), file = f)
        
def output_dataframe(df, f, sheetname):
    
    #Implement openpyxl workaround to prevent overwriting the whole file
    book = load_workbook(f)
    writer = pd.ExcelWriter(f, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    #Write new dataframe to Excel    
    df.to_excel(writer, sheet_name=sheetname)
    
    writer.save()

main()

if __name__ == "__main__" :
    main()